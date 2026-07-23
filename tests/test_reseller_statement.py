import os
import tempfile
import unittest
from datetime import date, datetime, timedelta


_DB_FILE = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
_DB_FILE.close()
os.environ['DATABASE_URL'] = f"sqlite:///{_DB_FILE.name.replace(os.sep, '/')}"
os.environ['FLASK_ENV'] = 'development'
os.environ['DISABLE_BACKGROUND_THREADS'] = '1'

from app import (Admin, ClientOwnership, Server, Transaction, UsageDaily,  # noqa: E402
                 app, db)

GB = 1024 ** 3


class ResellerStatementTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.ctx = app.app_context()
        cls.ctx.push()
        db.create_all()

    @classmethod
    def tearDownClass(cls):
        db.session.remove()
        db.drop_all()
        cls.ctx.pop()
        try:
            os.unlink(_DB_FILE.name)
        except OSError:
            pass

    def setUp(self):
        for model in (UsageDaily, ClientOwnership, Transaction, Server, Admin):
            model.query.delete()
        db.session.commit()

        self.owner = Admin(username='owner', password_hash='x',
                           role='superadmin', is_superadmin=True)
        self.reseller = Admin(username='res1', password_hash='x', role='reseller',
                              custom_cost_per_gb=2000)
        self.server = Server(name='srv-a', host='10.0.0.1',
                             username='u', password='p')
        db.session.add_all([self.owner, self.reseller, self.server])
        db.session.commit()

        now = datetime.utcnow()
        tx_purchase = Transaction(
            admin_id=self.reseller.id, server_id=self.server.id, amount=-100_000,
            type='purchase', category='usage', client_email='u1',
            package_name='P1', volume_gb=50, days=30,
            description='Purchase Package: P1 - u1', created_at=now - timedelta(hours=3))
        tx_renew = Transaction(
            admin_id=self.reseller.id, server_id=self.server.id, amount=-50_000,
            type='renew', category='usage', client_email='u1',
            package_name='P1', volume_gb=20, days=30,
            description='Renew Package: P1 - u1', created_at=now - timedelta(hours=2))
        tx_reset = Transaction(
            admin_id=self.reseller.id, server_id=self.server.id, amount=-10_000,
            type='reset_traffic', category='usage', client_email='u1',
            volume_gb=5, days=0, description='Reset traffic - u1',
            created_at=now - timedelta(hours=1))
        tx_deposit = Transaction(
            admin_id=self.reseller.id, amount=80_000, type='manual_receipt',
            category='income', description='Receipt', created_at=now)
        ownership = ClientOwnership(
            reseller_id=self.reseller.id, server_id=self.server.id,
            client_email='u1', client_uuid='sub-1')
        usage = UsageDaily(
            server_id=self.server.id, sub_id='sub-1', usage_date=date.today(),
            upload_bytes=5 * GB, download_bytes=5 * GB,
            opening_upload_bytes=0, opening_download_bytes=0,
            closing_upload_bytes=5 * GB, closing_download_bytes=5 * GB,
            sample_count=3)
        db.session.add_all([tx_purchase, tx_renew, tx_reset, tx_deposit,
                            ownership, usage])
        db.session.commit()

        self.client = app.test_client()

    def tearDown(self):
        db.session.remove()

    def _login(self, admin_id):
        with self.client.session_transaction() as sess:
            sess['admin_id'] = admin_id

    def _statement(self, extra=''):
        resp = self.client.get('/api/finance/reseller-statement?' + extra)
        self.assertEqual(resp.status_code, 200, resp.get_data(as_text=True))
        data = resp.get_json()
        self.assertTrue(data['success'])
        return data

    def test_summary_traffic_pricing_and_settlement(self):
        self._login(self.owner.id)
        data = self._statement(f'user_id={self.reseller.id}')
        s = data['summary']

        self.assertEqual(s['created'], 1)
        self.assertEqual(s['renewed'], 1)
        self.assertEqual(s['reset'], 1)
        self.assertEqual(s['spent'], 160_000)
        self.assertEqual(s['deposited'], 80_000)
        self.assertEqual(s['balance'], -80_000)
        self.assertEqual(s['revenue'], 80_000)
        self.assertEqual(s['should_pay'], 160_000)

        t = s['traffic']
        self.assertEqual(t['charged_gb'], 75)
        self.assertEqual(t['used_gb'], 10.0)
        self.assertEqual(t['remaining_gb'], 65.0)
        # weighted effective price: 160000 / 75 ≈ 2133 T/GB
        self.assertEqual(t['price_per_gb'], 2133)
        self.assertEqual(t['price_basis'], 'period_transactions')
        self.assertEqual(t['usage_cost'], 21330)
        self.assertEqual(t['accounts_tracked'], 1)

        self.assertEqual(s['pricing']['mode'], 'fixed')
        self.assertEqual(s['pricing']['custom_cost_per_gb'], 2000)
        self.assertNotIn('percent_settlement', s)

        # by_server rows now carry the server id for the lazy drill-down
        self.assertEqual(data['by_server'][0]['server_id'], self.server.id)

    def test_percent_settlement(self):
        self._login(self.owner.id)
        data = self._statement(f'user_id={self.reseller.id}&percent=50')
        ps = data['summary']['percent_settlement']
        self.assertEqual(ps['percent'], 50.0)
        self.assertEqual(ps['base'], 160_000)
        self.assertEqual(ps['debt'], 80_000)

    def test_accounts_breakdown_per_server(self):
        self._login(self.owner.id)
        resp = self.client.get(
            f'/api/finance/reseller-statement/accounts?user_id={self.reseller.id}'
            f'&server_id={self.server.id}')
        self.assertEqual(resp.status_code, 200, resp.get_data(as_text=True))
        accounts = resp.get_json()['accounts']
        self.assertEqual(len(accounts), 1)
        a = accounts[0]
        self.assertEqual(a['email'], 'u1')
        self.assertEqual(a['charged_gb'], 75)
        self.assertEqual(a['used_gb'], 10.0)
        self.assertEqual(a['remaining_gb'], 65.0)
        self.assertEqual(a['spent'], 160_000)
        self.assertEqual(a['events'], 3)
        self.assertEqual(a['last_type'], 'reset_traffic')
        self.assertEqual(a['status'], 'deleted')  # not present in the live cache

    def test_reseller_locked_to_own_statement(self):
        self._login(self.reseller.id)
        data = self._statement(f'user_id={self.owner.id}')  # ignored: not superadmin
        self.assertEqual(data['summary']['reseller']['id'], self.reseller.id)
        self.assertEqual(data['summary']['spent'], 160_000)

    def test_usage_resolved_via_live_subid(self):
        """Snapshots are keyed by the panel subId, which differs from the stored
        client uuid; the live cache must bridge them."""
        self._login(self.owner.id)
        usage = UsageDaily.query.first()
        usage.sub_id = 'sub-9'  # panel subId != stored uuid 'sub-1'
        db.session.commit()

        import app as app_module
        app_module.GLOBAL_SERVER_DATA['inbounds'] = [{
            'server_id': self.server.id, 'remark': 'inb',
            'clients': [{'id': 'sub-1', 'subId': 'sub-9', 'email': 'u1', 'enable': True}],
        }]
        try:
            data = self._statement(f'user_id={self.reseller.id}')
            t = data['summary']['traffic']
            self.assertEqual(t['used_gb'], 10.0)
            self.assertTrue(t['usage_available'])
            self.assertEqual(t['accounts_tracked'], 1)

            resp = self.client.get(
                f'/api/finance/reseller-statement/accounts?user_id={self.reseller.id}'
                f'&server_id={self.server.id}')
            accounts = resp.get_json()['accounts']
            self.assertEqual(len(accounts), 1)  # uuid + subId must not duplicate
            self.assertEqual(accounts[0]['status'], 'active')
            self.assertEqual(accounts[0]['used_gb'], 10.0)
        finally:
            app_module.GLOBAL_SERVER_DATA['inbounds'] = []

    def test_usage_warning_when_no_snapshots(self):
        self._login(self.owner.id)
        UsageDaily.query.delete()
        db.session.commit()
        data = self._statement(f'user_id={self.reseller.id}')
        t = data['summary']['traffic']
        self.assertFalse(t['usage_available'])
        self.assertIsNone(t['coverage_from_jalali'])
        self.assertEqual(t['used_gb'], 0)

    def test_export_xlsx(self):
        self._login(self.owner.id)
        resp = self.client.get(
            f'/api/finance/reseller-statement/export?user_id={self.reseller.id}&percent=50')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp.mimetype,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertTrue(resp.get_data().startswith(b'PK'))  # zip/xlsx magic

        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(resp.get_data()), read_only=True)
        self.assertEqual(wb.sheetnames, ['Summary', 'Daily', 'Accounts', 'Transactions'])
        wb.close()


if __name__ == '__main__':
    unittest.main()
