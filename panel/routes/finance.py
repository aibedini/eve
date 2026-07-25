"""Finance, payments, transactions, and reseller-statement API routes (extracted from app.py)."""
import io
import json
import re
from datetime import datetime, timedelta

from flask import Blueprint, jsonify, request, send_file, session
from jdatetime import datetime as jdatetime_class
from sqlalchemy import func, or_, text
from sqlalchemy.orm import joinedload

from panel.extensions import db, limiter
from panel.models import (
    Admin, BankCard, ClientOwnership, CustomerAccount, CustomerTransaction,
    ManualReceipt, Package, Payment, Server, Transaction, UsageDaily,
)
from panel.routes.common import login_required, user_management_required

bp = Blueprint('finance', __name__)


EMAIL_IN_DESCRIPTION = re.compile(r'([A-Za-z0-9_.+-]+@[A-Za-z0-9-]+\.[A-Za-z0-9-.]+)$')



def parse_amount_to_int(value):
    """Parse amount input to int.

    Accepts strings with commas/spaces and Persian/Arabic digits.
    Returns None if cannot parse.
    """
    from app import _DIGIT_TRANSLATION  # deferred: app-level helper, avoids circular import
    if value is None:
        return None
    try:
        if isinstance(value, (int, float)):
            return int(value)
        s = str(value).strip()
        if not s:
            return None
        s = s.translate(_DIGIT_TRANSLATION)
        # Keep digits only (strip separators/currency)
        s = re.sub(r'[^0-9]', '', s)
        if not s:
            return None
        return int(s)
    except Exception:
        return None



def extract_email_from_description(description):
    if not description:
        return None
    match = EMAIL_IN_DESCRIPTION.search(description.strip())
    if not match:
        return None
    email = match.group(1).strip().lower()
    return email.rstrip('.,;') or None



@bp.route('/admin/charge', methods=['POST'])
@user_management_required
def charge_admin():
    data = request.json
    admin_id = int(data.get('admin_id'))
    amount = int(data.get('amount'))
    description = data.get('description', 'Manual charge')
    
    admin = Admin.query.get_or_404(admin_id)
    admin.credit += amount
    
    transaction_type = 'deposit' if amount >= 0 else 'manual_debit'
    # Set category to 'income' for deposits (positive) so it counts in stats, 'expense' for debits (negative)
    category = 'income' if amount >= 0 else 'expense'

    transaction = Transaction(
        admin_id=admin_id,
        amount=amount,
        type=transaction_type,
        description=description,
        category=category
    )
    db.session.add(transaction)
    db.session.commit()
    return jsonify({"success": True, "new_credit": admin.credit})

@bp.route('/api/transactions', methods=['GET'])
@login_required
def get_transactions():
    from app import get_accessible_servers, parse_jalali_date  # deferred: app-level helper, avoids circular import
    try:
        user = db.session.get(Admin, session['admin_id'])
        if not user:
            return jsonify({"success": False, "error": "User not found"}), 401

        query = Transaction.query.join(Admin)

        type_filter = (request.args.get('type') or '').strip()
        direction_filter = (request.args.get('direction') or '').strip()

        if user.role == 'reseller':
            query = query.filter(Transaction.admin_id == user.id)
        else:
            target_user_id = request.args.get('user_id', type=int)
            if target_user_id:
                query = query.filter(Transaction.admin_id == target_user_id)

        # Filter by Server (using new column)
        server_filter = request.args.get('server_id', type=int)
        if server_filter:
            accessible_ids = {s.id for s in get_accessible_servers(user, include_disabled=True)}
            if user.role == 'reseller' and server_filter not in accessible_ids:
                return jsonify({"success": False, "error": "Access denied to requested server"}), 403
            query = query.filter(Transaction.server_id == server_filter)

        search_term = (request.args.get('search') or '').strip()
        if search_term:
            pattern = f"%{search_term}%"
            query = query.filter(or_(
                Transaction.client_email.ilike(pattern),
                Transaction.description.ilike(pattern),
                Transaction.type.ilike(pattern),
                Admin.username.ilike(pattern)
            ))

        if direction_filter == 'income':
            query = query.filter(Transaction.amount > 0)
        elif direction_filter == 'expense':
            query = query.filter(Transaction.amount < 0)

        if type_filter:
            query = query.filter(Transaction.type == type_filter)

        start_dt = parse_jalali_date(request.args.get('start_date'), end_of_day=False)
        if start_dt:
            query = query.filter(Transaction.created_at >= start_dt)

        end_dt = parse_jalali_date(request.args.get('end_date'), end_of_day=True)
        if end_dt:
            query = query.filter(Transaction.created_at <= end_dt)

        # ...existing code...
        # (rest of the function remains unchanged)
        # ...existing code...
    except Exception as ex:
        import traceback
        return jsonify({"success": False, "error": str(ex), "trace": traceback.format_exc()}), 500

    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('limit', 20, type=int)
    per_page = max(1, min(per_page, 100))

    pagination = query.order_by(Transaction.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    transactions = pagination.items

    # Fallback logic for old transactions (missing server_id)
    transaction_emails = {}
    email_pairs = set()
    for tx in transactions:
        if not tx.server_id:
            email = extract_email_from_description(tx.description)
            if email:
                transaction_emails[tx.id] = email
                email_pairs.add((tx.admin_id, email))

    ownership_map = {}
    if email_pairs:
        reseller_ids = {pair[0] for pair in email_pairs}
        email_values = {pair[1] for pair in email_pairs}
        if reseller_ids and email_values:
            ownerships = ClientOwnership.query.filter(
                ClientOwnership.reseller_id.in_(list(reseller_ids)),
                func.lower(ClientOwnership.client_email).in_(list(email_values))
            ).all()
            for ownership in ownerships:
                key = (ownership.reseller_id, (ownership.client_email or '').lower())
                existing = ownership_map.get(key)
                current_created = ownership.created_at or datetime.min
                existing_created = existing.created_at if existing and existing.created_at else datetime.min
                if not existing or current_created >= existing_created:
                    ownership_map[key] = ownership

    payload = []
    for tx in transactions:
        tx_data = tx.to_dict()
        
        # If server_id was missing, try to fill it from ownership map
        if not tx_data.get('server'):
            email = transaction_emails.get(tx.id)
            if email:
                tx_data['client_email'] = email
                ownership = ownership_map.get((tx.admin_id, email))
                if ownership and ownership.server:
                    tx_data['server_id'] = ownership.server.id
                    tx_data['server'] = {
                        'id': ownership.server.id,
                        'name': ownership.server.name
                    }

        payload.append(tx_data)

    return jsonify({
        'transactions': payload,
        'total': pagination.total,
        'pages': pagination.pages,
        'current_page': page,
        'per_page': per_page
    })


def _truncate_text(value: str, max_len: int) -> str:
    if value is None:
        return ''
    s = str(value)
    if len(s) <= max_len:
        return s
    return s[: max(0, max_len - 1)] + '…'


def _build_tx_edit_audit(editor_username: str, at_jalali: str, old_amount: int, old_type: str, old_category: str, old_desc: str, new_amount: int, new_type: str, new_category: str, new_desc: str) -> str:
    # Keep it compact (Transaction.description is VARCHAR(255))
    editor = editor_username or 'unknown'
    when = at_jalali or ''
    base = f"Edited by {editor} at {when}: {old_amount}->{new_amount}, {old_type}->{new_type}, {old_category}->{new_category}. "
    # Allocate remaining space for desc fragments
    remaining = 255 - len(base) - len("Was:  | Now: ")
    if remaining < 0:
        return _truncate_text(base, 255)
    half = max(0, remaining // 2)
    was_part = _truncate_text(old_desc or '', half)
    now_part = _truncate_text(new_desc or '', remaining - len(was_part))
    final = f"{base}Was: {was_part} | Now: {now_part}".strip()
    return _truncate_text(final, 255)


def _build_tx_delete_audit(deleter_username: str, at_jalali: str, deleted_tx_id: int, deleted_admin_username: str, deleted_amount: int, deleted_type: str) -> str:
    deleter = deleter_username or 'unknown'
    when = at_jalali or ''
    admin_u = deleted_admin_username or 'unknown'
    base = f"Deleted by {deleter} at {when}: tx#{deleted_tx_id} ({admin_u}) {deleted_amount} {deleted_type}."
    return _truncate_text(base, 255)


@bp.route('/api/transactions/<int:tx_id>', methods=['PUT'])
@user_management_required
def update_transaction(tx_id):
    from app import format_jalali, parse_jalali_date  # deferred: app-level helper, avoids circular import
    editor = db.session.get(Admin, session.get('admin_id'))
    tx = Transaction.query.get_or_404(tx_id)

    try:
        data = request.get_json() or {}
    except Exception:
        return jsonify({"success": False, "error": "Invalid JSON"}), 400

    # Snapshot for audit
    old_amount = tx.amount
    old_type = tx.type or ''
    old_category = tx.category or ''
    old_desc = tx.description or ''

    # Determine direction
    is_expense = None
    if 'is_expense' in data:
        is_expense = bool(data.get('is_expense'))

    # Amount (UI sends positive digits; we apply sign based on direction)
    parsed_amount = None
    if 'amount' in data:
        parsed_amount = parse_amount_to_int(data.get('amount'))
        if parsed_amount is None or int(parsed_amount) <= 0:
            return jsonify({"success": False, "error": "Invalid amount"}), 400

    if is_expense is None:
        # infer from existing
        is_expense = (tx.amount or 0) < 0 or (tx.category == 'expense')

    if parsed_amount is not None:
        tx.amount = -abs(int(parsed_amount)) if is_expense else abs(int(parsed_amount))

    # Category
    tx.category = 'expense' if is_expense else 'income'

    # Type
    new_type = (data.get('cost_type') or data.get('type') or tx.type or '').strip() or None
    if new_type is not None:
        tx.type = new_type

    # Common fields
    if 'server_id' in data:
        tx.server_id = data.get('server_id') or None
    if 'card_id' in data:
        tx.card_id = data.get('card_id') or None
    if 'sender_card' in data:
        tx.sender_card = (data.get('sender_card') or '').strip() or None
    if 'sender_name' in data:
        tx.sender_name = (data.get('sender_name') or '').strip() or None
    if 'client_email' in data:
        tx.client_email = (data.get('client_email') or '').strip() or None

    # Date/time (Jalali + Tehran)
    if 'payment_date' in data or 'payment_time' in data:
        date_part = (data.get('payment_date') or '').strip() or None
        time_part = (data.get('payment_time') or '').strip() or None
        combined = None
        if date_part and time_part:
            combined = f"{date_part} {time_part}"
        elif date_part:
            combined = date_part
        if combined:
            dt = parse_jalali_date(combined, end_of_day=False)
            if dt:
                tx.created_at = dt

    # Description update + audit
    new_desc = old_desc
    if 'description' in data:
        new_desc = (data.get('description') or '').strip()

    audit_desc = _build_tx_edit_audit(
        editor.username if editor else 'unknown',
        format_jalali(datetime.utcnow()) or '',
        old_amount,
        old_type,
        old_category,
        old_desc,
        tx.amount,
        (tx.type or ''),
        (tx.category or ''),
        new_desc,
    )
    tx.description = audit_desc

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"success": False, "error": "Failed to update transaction"}), 500

    return jsonify({"success": True, "transaction": tx.to_dict()})


@bp.route('/api/transactions/<int:tx_id>', methods=['DELETE'])
@user_management_required
def delete_transaction(tx_id):
    from app import format_jalali  # deferred: app-level helper, avoids circular import
    deleter = db.session.get(Admin, session.get('admin_id'))
    tx = Transaction.query.get_or_404(tx_id)

    # Create an audit log entry that remains visible in /transactions
    try:
        deleted_admin_username = None
        if hasattr(tx, 'admin') and tx.admin:
            deleted_admin_username = tx.admin.username
        else:
            admin_obj = db.session.get(Admin, tx.admin_id)
            deleted_admin_username = admin_obj.username if admin_obj else None

        audit_desc = _build_tx_delete_audit(
            deleter.username if deleter else 'unknown',
            format_jalali(datetime.utcnow()) or '',
            tx.id,
            deleted_admin_username,
            int(tx.amount or 0),
            (tx.type or ''),
        )

        audit_tx = Transaction(
            admin_id=deleter.id if deleter else tx.admin_id,
            amount=0,
            type='audit',
            category='usage',
            description=audit_desc,
            created_at=datetime.utcnow(),
        )
        db.session.add(audit_tx)
    except Exception:
        # If audit creation fails, continue with delete (avoid blocking admin cleanup)
        pass

    db.session.delete(tx)
    db.session.commit()
    return jsonify({"success": True})


# ==================== FINANCE OVERVIEW ====================



@bp.route('/api/payments', methods=['GET'])
@login_required
def get_payments():
    """Get payment transactions (transactions that have card info)"""
    from app import format_jalali, parse_jalali_date  # deferred: app-level helper, avoids circular import
    try:
        user = db.session.get(Admin, session['admin_id'])
        if not user:
            return jsonify({"success": False, "error": "User not found"}), 401
        
        type_filter = (request.args.get('type') or '').strip()
        direction_filter = (request.args.get('direction') or '').strip()

        # Payments
        payment_query = Payment.query
        if user.role == 'reseller':
            payment_query = payment_query.filter(Payment.admin_id == user.id)
        else:
            target_user_id = request.args.get('user_id', type=int)
            if target_user_id:
                payment_query = payment_query.filter(Payment.admin_id == target_user_id)
        card_id = request.args.get('card_id', type=int)
        if card_id:
            payment_query = payment_query.filter(Payment.card_id == card_id)
        search_term = (request.args.get('search') or '').strip()
        if search_term:
            pattern = f"%{search_term}%"
            payment_query = payment_query.filter(or_(
                Payment.description.ilike(pattern),
                Payment.sender_card.ilike(pattern),
                Payment.sender_name.ilike(pattern),
                Payment.client_email.ilike(pattern)
            ))
        start_dt = parse_jalali_date(request.args.get('start_date'), end_of_day=False)
        if start_dt:
            payment_query = payment_query.filter(Payment.payment_date >= start_dt)
        end_dt = parse_jalali_date(request.args.get('end_date'), end_of_day=True)
        if end_dt:
            payment_query = payment_query.filter(Payment.payment_date <= end_dt)
        include_payments = (direction_filter != 'expense' and (not type_filter or type_filter == 'payment'))

        # ...existing code...
        # (rest of the function remains unchanged)
        # ...existing code...
    except Exception as ex:
        import traceback
        return jsonify({"success": False, "error": str(ex), "trace": traceback.format_exc()}), 500
    tx_query = Transaction.query
    if user.role == 'reseller':
        tx_query = tx_query.filter(Transaction.admin_id == user.id)
    else:
        target_user_id = request.args.get('user_id', type=int)
        if target_user_id:
            tx_query = tx_query.filter(Transaction.admin_id == target_user_id)
    if card_id:
        tx_query = tx_query.filter(Transaction.card_id == card_id)
    server_id = request.args.get('server_id', type=int)
    if server_id:
        tx_query = tx_query.filter(Transaction.server_id == server_id)
    if search_term:
        pattern = f"%{search_term}%"
        tx_query = tx_query.filter(or_(
            Transaction.description.ilike(pattern),
            Transaction.sender_card.ilike(pattern),
            Transaction.sender_name.ilike(pattern),
            Transaction.client_email.ilike(pattern),
            Transaction.type.ilike(pattern)
        ))

    # Exclude system/audit rows from Finance overview list
    tx_query = tx_query.filter(Transaction.type != 'audit')

    if direction_filter == 'income':
        tx_query = tx_query.filter(Transaction.amount > 0)
    elif direction_filter == 'expense':
        tx_query = tx_query.filter(Transaction.amount < 0)

    if type_filter and type_filter not in ('payment', 'receipt'):
        tx_query = tx_query.filter(Transaction.type == type_filter)
    elif type_filter in ('payment', 'receipt'):
        tx_query = tx_query.filter(text('1=0'))
    if start_dt:
        tx_query = tx_query.filter(Transaction.created_at >= start_dt)
    if end_dt:
        tx_query = tx_query.filter(Transaction.created_at <= end_dt)
    include_transactions = True

    # ManualReceipts
    receipt_query = ManualReceipt.query
    if user.role == 'reseller':
        receipt_query = receipt_query.filter(ManualReceipt.admin_id == user.id)
    else:
        target_user_id = request.args.get('user_id', type=int)
        if target_user_id:
            receipt_query = receipt_query.filter(ManualReceipt.admin_id == target_user_id)
    if card_id:
        receipt_query = receipt_query.filter(ManualReceipt.card_id == card_id)
    if search_term:
        pattern = f"%{search_term}%"
        receipt_query = receipt_query.filter(or_(
            ManualReceipt.reference_code.ilike(pattern),
            ManualReceipt.notes.ilike(pattern)
        ))
    if start_dt:
        receipt_query = receipt_query.filter(ManualReceipt.deposit_at >= start_dt)
    if end_dt:
        receipt_query = receipt_query.filter(ManualReceipt.deposit_at <= end_dt)
    include_receipts = (direction_filter != 'expense' and (not type_filter or type_filter == 'receipt'))

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('limit', 20, type=int)
    page = max(1, int(page or 1))
    per_page = max(1, min(int(per_page or 20), 100))

    payment_count = payment_query.order_by(None).count() if include_payments else 0
    tx_count = tx_query.order_by(None).count() if include_transactions else 0
    receipt_count = receipt_query.order_by(None).count() if include_receipts else 0

    total = int(payment_count) + int(tx_count) + int(receipt_count)
    pages = (total + per_page - 1) // per_page if total > 0 else 1

    start = (page - 1) * per_page
    end = start + per_page
    fetch_limit = max(end, per_page)

    payments_list = []
    if include_payments and payment_count > 0:
        payments_list = payment_query.options(
            joinedload(Payment.admin),
            joinedload(Payment.card),
        ).order_by(Payment.payment_date.desc()).limit(fetch_limit).all()

    transactions_list = []
    if include_transactions and tx_count > 0:
        transactions_list = tx_query.options(
            joinedload(Transaction.admin),
            joinedload(Transaction.card),
            joinedload(Transaction.server),
        ).order_by(Transaction.created_at.desc()).limit(fetch_limit).all()

    receipts_list = []
    if include_receipts and receipt_count > 0:
        receipts_list = receipt_query.options(
            joinedload(ManualReceipt.admin),
            joinedload(ManualReceipt.card),
        ).order_by(ManualReceipt.deposit_at.desc()).limit(fetch_limit).all()

    # Map payments
    mapped_payments = []
    for p in payments_list:
        d = p.to_dict()
        d['type'] = 'payment'
        mapped_payments.append(d)

    # Map transactions
    mapped_transactions = []
    for t in transactions_list:
        admin = getattr(t, 'admin', None)
        card = getattr(t, 'card', None)
        server = getattr(t, 'server', None)
        jalali_date = format_jalali(t.created_at) or ''
        mapped_transactions.append({
            'id': f"tx-{t.id}",
            'admin_id': t.admin_id,
            'admin': {
                'id': admin.id,
                'username': admin.username,
                'role': admin.role
            } if admin else None,
            'sender_card': t.sender_card or '',
            'sender_name': getattr(t, 'sender_name', None) or None,
            'card_id': t.card_id,
            'card': {
                'id': card.id,
                'label': card.label,
                'bank_name': card.bank_name
            } if card else None,
            'server': {
                'id': server.id,
                'name': server.name
            } if server else None,
            'amount': int(t.amount),
            'type': t.type or 'transaction',
            'description': t.description,
            'client_email': t.client_email or (t.description.split(' - ')[-1] if t.description and ' - ' in t.description else ''),
            'payment_date': t.created_at.isoformat() if t.created_at else None,
            'payment_date_jalali': jalali_date,
            'verified': True,
            'created_at': t.created_at.isoformat() if t.created_at else None,
        })

    # Map receipts
    mapped_receipts = []
    for r in receipts_list:
        d = r.to_dict()
        d['type'] = 'receipt'
        d['payment_date'] = d.get('deposit_at') or d.get('created_at')
        d['payment_date_jalali'] = ''
        if d['payment_date']:
            try:
                dt = datetime.fromisoformat(d['payment_date'])
                d['payment_date_jalali'] = format_jalali(dt)
            except:
                pass
        mapped_receipts.append(d)

    # Combine all
    combined = mapped_payments + mapped_transactions + mapped_receipts
    def get_date(item):
        date_str = item.get('payment_date') or item.get('created_at')
        if not date_str:
            return datetime.min
        try:
            return datetime.fromisoformat(date_str)
        except:
            try:
                return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            except:
                return datetime.min
    combined.sort(key=lambda x: get_date(x), reverse=True)
    page_items = combined[start:end]
    return jsonify({
        'payments': page_items,
        'total': total,
        'pages': pages,
        'current_page': page,
        'per_page': per_page
    })


def _finance_request_filters() -> dict:
    from app import parse_jalali_date  # deferred: app-level helper, avoids circular import
    return {
        'type': (request.args.get('type') or '').strip(),
        'direction': (request.args.get('direction') or '').strip(),
        'search': (request.args.get('search') or '').strip(),
        'start_dt': parse_jalali_date(request.args.get('start_date'), end_of_day=False),
        'end_dt': parse_jalali_date(request.args.get('end_date'), end_of_day=True),
    }


def _empty_query(q):
    return q.filter(text('1=0'))


def _apply_finance_tx_filters(q, filters: dict, *, date_column=None,
                              allow_income=True, allow_expense=True):
    """Apply the same user-visible Finance filters used by the list endpoint.

    Overview/stat cards must not silently include rows that the current Finance
    list has filtered out; otherwise an edit can appear in totals while being
    invisible in the table.
    """
    date_column = date_column or Transaction.created_at
    direction_filter = filters.get('direction') or ''
    type_filter = filters.get('type') or ''
    search_term = filters.get('search') or ''

    if direction_filter == 'income' and not allow_income:
        return _empty_query(q)
    if direction_filter == 'expense' and not allow_expense:
        return _empty_query(q)

    if type_filter:
        if type_filter in ('payment', 'receipt'):
            return _empty_query(q)
        q = q.filter(Transaction.type == type_filter)

    if search_term:
        pattern = f"%{search_term}%"
        q = q.filter(or_(
            Transaction.description.ilike(pattern),
            Transaction.sender_card.ilike(pattern),
            Transaction.sender_name.ilike(pattern),
            Transaction.client_email.ilike(pattern),
            Transaction.type.ilike(pattern)
        ))

    if filters.get('start_dt'):
        q = q.filter(date_column >= filters['start_dt'])
    if filters.get('end_dt'):
        q = q.filter(date_column <= filters['end_dt'])
    return q


def _apply_finance_payment_filters(q, filters: dict, *, date_column=None):
    date_column = date_column or Payment.payment_date
    direction_filter = filters.get('direction') or ''
    type_filter = filters.get('type') or ''
    search_term = filters.get('search') or ''

    if direction_filter == 'expense':
        return _empty_query(q)
    if type_filter and type_filter != 'payment':
        return _empty_query(q)

    if search_term:
        pattern = f"%{search_term}%"
        q = q.filter(or_(
            Payment.description.ilike(pattern),
            Payment.sender_card.ilike(pattern),
            Payment.sender_name.ilike(pattern),
            Payment.client_email.ilike(pattern)
        ))

    if filters.get('start_dt'):
        q = q.filter(date_column >= filters['start_dt'])
    if filters.get('end_dt'):
        q = q.filter(date_column <= filters['end_dt'])
    return q


@bp.route('/api/payments', methods=['POST'])
@login_required
def add_payment():
    from app import logger, parse_jalali_date  # deferred: app-level helper, avoids circular import
    user = db.session.get(Admin, session['admin_id'])
    if not user:
        return jsonify({"success": False, "error": "User not found"}), 401
    
    try:
        data = request.get_json() or {}
    except Exception as ex:
        logger.exception('Invalid JSON in add_payment')
        return jsonify({"success": False, "error": "Invalid JSON"}), 400

    # Log incoming request for debugging
    try:
        logger.info('add_payment request by user %s: %s', user.id, json.dumps(data, ensure_ascii=False))
    except Exception:
        logger.info('add_payment request: (unserializable data)')
    
    amount_val = parse_amount_to_int(data.get('amount'))
    if not amount_val or int(amount_val) <= 0:
        return jsonify({"success": False, "error": "Amount is required and must be positive"}), 400
    
    payment_date_str = (data.get('payment_date') or '').strip() or None
    payment_time_str = (data.get('payment_time') or '').strip() or None
    combined_dt_str = None
    if payment_date_str and payment_time_str:
        combined_dt_str = f"{payment_date_str} {payment_time_str}"
    elif payment_date_str:
        combined_dt_str = payment_date_str

    # parse_jalali_date converts Tehran -> UTC; ensure we pass date+time together to avoid double shifting
    payment_date = parse_jalali_date(combined_dt_str, end_of_day=False)
    if not payment_date:
        payment_date = datetime.utcnow()
    
    # Expense flow: create a Transaction with category='expense' (amount negative)
    is_expense = bool(data.get('is_expense'))
    if is_expense:
        cost_type = (data.get('cost_type') or 'server_cost').strip()
        server_id = data.get('server_id') or None
        amount_val = -abs(int(amount_val))

        tx = Transaction(
            admin_id=user.id,
            server_id=server_id,
            card_id=data.get('card_id') or None,
            sender_card=(data.get('sender_card') or '').strip() or None,
            sender_name=(data.get('sender_name') or '').strip() or None,
            client_email=(data.get('client_email') or '').strip() or None,
            amount=amount_val,
            type=cost_type,
            description=data.get('description', '').strip() or None,
            category='expense',
            created_at=payment_date
        )
        db.session.add(tx)
        try:
            db.session.commit()
            logger.info('Expense saved as transaction id=%s admin=%s amount=%s server=%s type=%s', tx.id, tx.admin_id, tx.amount, server_id, cost_type)
        except Exception:
            logger.exception('Failed to commit expense transaction')
            db.session.rollback()
            return jsonify({"success": False, "error": "Failed to save expense"}), 500
        return jsonify({"success": True, "id": tx.id, "mode": "expense"})

    # Income (payment) flow
    is_super = (user.role == 'superadmin' or user.is_superadmin)
    payment = Payment(
        admin_id=user.id,
        card_id=data.get('card_id') or None,
        sender_card=data.get('sender_card', '').strip() or None,
        sender_name=data.get('sender_name', '').strip() or None,
        amount=int(amount_val),
        payment_date=payment_date,
        client_email=data.get('client_email', '').strip() or None,
        description=data.get('description', '').strip() or None,
        verified=True if is_super else False
    )
    
    db.session.add(payment)
    try:
        db.session.commit()
        logger.info('Payment saved id=%s admin=%s amount=%s', payment.id, payment.admin_id, payment.amount)
    except Exception as e:
        logger.exception('Failed to commit payment')
        db.session.rollback()
        return jsonify({"success": False, "error": "Failed to save payment"}), 500
    
    return jsonify({"success": True, "payment": payment.to_dict()})


@bp.route('/api/payments/<int:payment_id>', methods=['PUT'])
@login_required
def update_payment(payment_id):
    from app import format_jalali, parse_jalali_date  # deferred: app-level helper, avoids circular import
    user = db.session.get(Admin, session['admin_id'])
    if not user:
        return jsonify({"success": False, "error": "User not found"}), 401
    
    payment = Payment.query.get_or_404(payment_id)
    
    is_super = (user.role == 'superadmin' or user.is_superadmin)
    # Only owner or superadmin can edit
    if payment.admin_id != user.id and not is_super:
        return jsonify({"success": False, "error": "Access denied"}), 403
    
    try:
        data = request.get_json() or {}
    except:
        return jsonify({"success": False, "error": "Invalid JSON"}), 400

    # Convert payment -> expense transaction (when editing and user switches kind)
    if bool(data.get('is_expense')):
        parsed_amount = parse_amount_to_int(data.get('amount')) if 'amount' in data else payment.amount
        if parsed_amount is None or int(parsed_amount) <= 0:
            return jsonify({"success": False, "error": "Invalid amount"}), 400

        server_id = data.get('server_id') or None
        if not server_id:
            return jsonify({"success": False, "error": "Server is required for expense"}), 400

        # Resolve date/time (Jalali + Tehran)
        resolved_dt = payment.payment_date
        if 'payment_date' in data or 'payment_time' in data:
            date_part = (data.get('payment_date') or '').strip() or None
            time_part = (data.get('payment_time') or '').strip() or None

            if not date_part and time_part and payment.payment_date:
                # derive current Jalali date part from existing UTC datetime
                try:
                    current_tehran = payment.payment_date + timedelta(hours=3, minutes=30)
                    j_current = jdatetime_class.fromgregorian(datetime=current_tehran)
                    date_part = j_current.strftime('%Y/%m/%d')
                except Exception:
                    date_part = None

            combined_dt_str = None
            if date_part and time_part:
                combined_dt_str = f"{date_part} {time_part}"
            elif date_part:
                combined_dt_str = date_part

            new_date = parse_jalali_date(combined_dt_str, end_of_day=False)
            if new_date:
                resolved_dt = new_date

        # Carry forward fields (prefer explicit updates)
        card_id = data.get('card_id') if 'card_id' in data else payment.card_id
        sender_card = (data.get('sender_card') if 'sender_card' in data else payment.sender_card) or None
        sender_name = (data.get('sender_name') if 'sender_name' in data else payment.sender_name) or None
        client_email = (data.get('client_email') if 'client_email' in data else payment.client_email) or None
        base_desc = (data.get('description') if 'description' in data else payment.description) or ''

        sender_card = (sender_card or '').strip() or None
        sender_name = (sender_name or '').strip() or None
        client_email = (client_email or '').strip() or None
        base_desc = (base_desc or '').strip()

        cost_type = (data.get('cost_type') or 'server_cost').strip() or 'server_cost'

        audit_note = _truncate_text(
            f"Converted from payment#{payment.id} by {(user.username if user else 'unknown')} at {format_jalali(datetime.utcnow()) or ''}.",
            255,
        )
        merged_desc = (f"{base_desc} {audit_note}".strip())
        merged_desc = _truncate_text(merged_desc, 255) or None

        tx = Transaction(
            admin_id=payment.admin_id,
            server_id=server_id,
            card_id=card_id or None,
            sender_card=sender_card,
            sender_name=sender_name,
            client_email=client_email,
            amount=-abs(int(parsed_amount)),
            type=cost_type,
            description=merged_desc,
            category='expense',
            created_at=resolved_dt or datetime.utcnow(),
        )

        db.session.add(tx)
        db.session.delete(payment)
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            return jsonify({"success": False, "error": "Failed to convert payment to expense"}), 500

        return jsonify({
            "success": True,
            "converted": True,
            "mode": "expense",
            "transaction_id": tx.id,
            "entry_id": f"tx-{tx.id}",
        })
    
    if 'amount' in data:
        parsed = parse_amount_to_int(data.get('amount'))
        if parsed is None or int(parsed) <= 0:
            return jsonify({"success": False, "error": "Invalid amount"}), 400
        payment.amount = int(parsed)
    if 'card_id' in data:
        payment.card_id = data['card_id'] or None
    if 'sender_card' in data:
        payment.sender_card = data['sender_card'].strip() or None
    if 'sender_name' in data:
        payment.sender_name = data['sender_name'].strip() or None
    if 'client_email' in data:
        payment.client_email = data['client_email'].strip() or None
    if 'description' in data:
        payment.description = data['description'].strip() or None
    if 'verified' in data and is_super:
        payment.verified = bool(data['verified'])
    if 'payment_date' in data or 'payment_time' in data:
        date_part = (data.get('payment_date') or '').strip() or None
        time_part = (data.get('payment_time') or '').strip() or None

        if not date_part and time_part and payment.payment_date:
            # derive current Jalali date part from existing UTC datetime
            try:
                current_tehran = payment.payment_date + timedelta(hours=3, minutes=30)
                j_current = jdatetime_class.fromgregorian(datetime=current_tehran)
                date_part = j_current.strftime('%Y/%m/%d')
            except Exception:
                date_part = None

        combined_dt_str = None
        if date_part and time_part:
            combined_dt_str = f"{date_part} {time_part}"
        elif date_part:
            combined_dt_str = date_part

        new_date = parse_jalali_date(combined_dt_str, end_of_day=False)
        if new_date:
            payment.payment_date = new_date
    
    db.session.commit()
    return jsonify({"success": True, "payment": payment.to_dict()})


@bp.route('/api/payments/<int:payment_id>', methods=['DELETE'])
@login_required
def delete_payment(payment_id):
    user = db.session.get(Admin, session['admin_id'])
    if not user:
        return jsonify({"success": False, "error": "User not found"}), 401
    
    payment = Payment.query.get_or_404(payment_id)
    
    is_super = (user.role == 'superadmin' or user.is_superadmin)
    # Only owner or superadmin can delete
    if payment.admin_id != user.id and not is_super:
        return jsonify({"success": False, "error": "Access denied"}), 403
    
    db.session.delete(payment)
    db.session.commit()
    return jsonify({"success": True})


@bp.route('/api/finance/stats', methods=['GET'])
@login_required
def get_finance_stats():
    """Get income statistics from transactions: today, this week, this month"""
    user = db.session.get(Admin, session['admin_id'])
    if not user:
        return jsonify({"success": False, "error": "User not found"}), 401
    
    # Calculate dates based on Tehran time and Jalali calendar
    now_utc = datetime.utcnow()
    tehran_offset = timedelta(hours=3, minutes=30)
    now_tehran = now_utc + tehran_offset
    
    j_now = jdatetime_class.fromgregorian(datetime=now_tehran)
    
    # Start of Jalali month
    j_month_start = j_now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    g_month_start_tehran = j_month_start.togregorian()
    month_start = g_month_start_tehran - tehran_offset

    # Start of previous Jalali month
    try:
        prev_year = int(j_month_start.year)
        prev_month = int(j_month_start.month) - 1
        if prev_month <= 0:
            prev_month = 12
            prev_year -= 1
        j_prev_month_start = jdatetime_class(prev_year, prev_month, 1, 0, 0, 0, 0)
        g_prev_month_start_tehran = j_prev_month_start.togregorian()
        prev_month_start = g_prev_month_start_tehran - tehran_offset
        prev_month_end = month_start
    except Exception:
        prev_month_start = None
        prev_month_end = None
    
    # Start of Today (Jalali/Tehran)
    j_today_start = j_now.replace(hour=0, minute=0, second=0, microsecond=0)
    g_today_start_tehran = j_today_start.togregorian()
    today_start = g_today_start_tehran - tehran_offset

    # Period-to-date windows for "compare to same period last month"
    now_period_end = now_utc
    try:
        cur_period_len = now_period_end - month_start
    except Exception:
        cur_period_len = None

    prev_period_end = None
    if prev_month_start and cur_period_len is not None:
        try:
            prev_period_end = prev_month_start + cur_period_len
            if prev_month_end and prev_period_end > prev_month_end:
                prev_period_end = prev_month_end
        except Exception:
            prev_period_end = prev_month_end

    # "Today" comparison: same day-of-month in previous Jalali month, same time-of-day window
    prev_same_day_start = None
    prev_same_day_end = None
    if prev_month_start:
        try:
            day_of_month = int(getattr(j_now, 'day', 0) or 0)
        except Exception:
            day_of_month = 0

        if day_of_month > 0:
            try:
                prev_year = int(j_month_start.year)
                prev_month = int(j_month_start.month) - 1
                if prev_month <= 0:
                    prev_month = 12
                    prev_year -= 1

                # Handle months that don't have this day (e.g., day 31)
                safe_day = day_of_month
                j_prev_same_day = None
                while safe_day >= 1 and j_prev_same_day is None:
                    try:
                        j_prev_same_day = jdatetime_class(prev_year, prev_month, safe_day, 0, 0, 0, 0)
                    except Exception:
                        safe_day -= 1

                if j_prev_same_day is not None:
                    g_prev_same_day_tehran = j_prev_same_day.togregorian()
                    prev_same_day_start = g_prev_same_day_tehran - tehran_offset

                    # Match window length to "today so far" but clamp to that day's end
                    try:
                        today_so_far = now_period_end - today_start
                    except Exception:
                        today_so_far = timedelta(0)
                    day_end = prev_same_day_start + timedelta(days=1)
                    prev_same_day_end = prev_same_day_start + today_so_far
                    if prev_same_day_end > day_end:
                        prev_same_day_end = day_end
            except Exception:
                prev_same_day_start = None
                prev_same_day_end = None

    card_id = request.args.get('card_id', type=int)
    server_id = request.args.get('server_id', type=int)
    target_user_id = request.args.get('user_id', type=int)
    finance_filters = _finance_request_filters()

    excluded_types_upper = {'SERVER_COST', 'SERVER_RENEWAL', 'SERVER_TRAFFIC'}

    if user.role == 'reseller':
        # Charge: مجموع تراکنش‌های مثبت (واریزها)
        charge_query = Transaction.query.filter(
            Transaction.admin_id == user.id,
            Transaction.amount > 0,
            or_(Transaction.type.is_(None), func.upper(Transaction.type).notin_(excluded_types_upper))
        )
        usage_query = Transaction.query.filter(Transaction.admin_id == user.id, Transaction.amount < 0)
        if card_id:
            charge_query = charge_query.filter(Transaction.card_id == card_id)
            usage_query = usage_query.filter(Transaction.card_id == card_id)
        if server_id:
            charge_query = charge_query.filter(Transaction.server_id == server_id)
            usage_query = usage_query.filter(Transaction.server_id == server_id)
        charge_query = _apply_finance_tx_filters(charge_query, finance_filters, allow_expense=False)
        usage_query = _apply_finance_tx_filters(usage_query, finance_filters, allow_income=False)

        def sum_amount(q, start_time=None):
            if start_time:
                q = q.filter(Transaction.created_at >= start_time)
            return db.session.query(func.coalesce(func.sum(Transaction.amount), 0)).filter(Transaction.id.in_(q.with_entities(Transaction.id))).scalar() or 0

        today_charge = sum_amount(charge_query, today_start)
        month_charge = sum_amount(charge_query, month_start)
        total_charge = sum_amount(charge_query)

        prev_month_charge = 0
        if prev_month_start and prev_period_end:
            q_prev = charge_query.filter(Transaction.created_at >= prev_month_start, Transaction.created_at < prev_period_end)
            prev_month_charge = db.session.query(func.coalesce(func.sum(Transaction.amount), 0)).filter(Transaction.id.in_(q_prev.with_entities(Transaction.id))).scalar() or 0

        prev_same_day_charge = 0
        if prev_same_day_start and prev_same_day_end:
            q_prev_day = charge_query.filter(Transaction.created_at >= prev_same_day_start, Transaction.created_at < prev_same_day_end)
            prev_same_day_charge = db.session.query(func.coalesce(func.sum(Transaction.amount), 0)).filter(Transaction.id.in_(q_prev_day.with_entities(Transaction.id))).scalar() or 0

        month_usage = abs(sum_amount(usage_query, month_start))
        total_usage = abs(sum_amount(usage_query))

        prev_month_usage = 0
        if prev_month_start and prev_period_end:
            q_prev_u = usage_query.filter(Transaction.created_at >= prev_month_start, Transaction.created_at < prev_period_end)
            prev_month_usage = abs(db.session.query(func.coalesce(func.sum(Transaction.amount), 0)).filter(Transaction.id.in_(q_prev_u.with_entities(Transaction.id))).scalar() or 0)

        month_net = int(month_charge) - int(month_usage)
        prev_month_net = int(prev_month_charge) - int(prev_month_usage)

        def pct_change(cur, prev):
            try:
                cur = float(cur)
                prev = float(prev)
            except Exception:
                return None
            if prev == 0:
                return 0.0 if cur == 0 else None
            return ((cur - prev) / abs(prev)) * 100.0

        month_charge_pct = pct_change(month_charge, prev_month_charge)
        month_usage_pct = pct_change(month_usage, prev_month_usage)
        month_net_pct = pct_change(month_net, prev_month_net)

        today_charge_pct = pct_change(today_charge, prev_same_day_charge)

        month_charge_delta = int(month_charge) - int(prev_month_charge)
        month_usage_delta = int(month_usage) - int(prev_month_usage)
        month_net_delta = int(month_net) - int(prev_month_net)
        today_charge_delta = int(today_charge) - int(prev_same_day_charge)

        remain = total_charge - total_usage

        return jsonify({
            'success': True,
            'stats': {
                'today': today_charge,
                'month': month_charge,
                'month_expense': month_usage,
                'total': remain,
                'prev_month': prev_month_charge,
                'prev_month_expense': prev_month_usage,
                'month_net': month_net,
                'prev_month_net': prev_month_net,
                'month_change_pct': month_charge_pct,
                'month_expense_change_pct': month_usage_pct,
                'month_net_change_pct': month_net_pct,
                'today_prev': prev_same_day_charge,
                'today_change_pct': today_charge_pct,
                'today_change_amount': today_charge_delta,
                'month_change_amount': month_charge_delta,
                'month_expense_change_amount': month_usage_delta,
                'month_net_change_amount': month_net_delta,
                'payment_count': charge_query.count(),
                'total_charge': total_charge,
                'total_usage': total_usage
            }
        })
    else:
        # Month income must include ALL positive entries except server cost/renewal/traffic.
        # Also include manual `Payment` records (these are separate from `Transaction`).
        tx_income_query = Transaction.query.filter(
            Transaction.amount > 0,
            or_(Transaction.type.is_(None), func.upper(Transaction.type).notin_(excluded_types_upper))
        )
        if target_user_id:
            tx_income_query = tx_income_query.filter(Transaction.admin_id == target_user_id)
        if card_id:
            tx_income_query = tx_income_query.filter(Transaction.card_id == card_id)
        if server_id:
            tx_income_query = tx_income_query.filter(Transaction.server_id == server_id)
        tx_income_query = _apply_finance_tx_filters(tx_income_query, finance_filters, allow_expense=False)

        pay_income_query = Payment.query.filter(Payment.amount > 0)
        if target_user_id:
            pay_income_query = pay_income_query.filter(Payment.admin_id == target_user_id)
        if card_id:
            pay_income_query = pay_income_query.filter(Payment.card_id == card_id)
        pay_income_query = _apply_finance_payment_filters(pay_income_query, finance_filters)

        def sum_tx_income(start_time=None, end_time=None):
            q = db.session.query(func.coalesce(func.sum(Transaction.amount), 0)).filter(
                Transaction.id.in_(tx_income_query.with_entities(Transaction.id))
            )
            if start_time:
                q = q.filter(Transaction.created_at >= start_time)
            if end_time:
                q = q.filter(Transaction.created_at < end_time)
            return q.scalar() or 0

        def sum_pay_income(start_time=None, end_time=None):
            q = db.session.query(func.coalesce(func.sum(Payment.amount), 0)).filter(
                Payment.id.in_(pay_income_query.with_entities(Payment.id))
            )
            if start_time:
                q = q.filter(Payment.payment_date >= start_time)
            if end_time:
                q = q.filter(Payment.payment_date < end_time)
            return q.scalar() or 0

        today_income = sum_tx_income(today_start) + sum_pay_income(today_start)
        month_income = sum_tx_income(month_start) + sum_pay_income(month_start)
        total_income = sum_tx_income() + sum_pay_income()

        prev_month_income = 0
        if prev_month_start and prev_period_end:
            prev_month_income = (
                sum_tx_income(prev_month_start, prev_period_end)
                + sum_pay_income(prev_month_start, prev_period_end)
            )

        prev_same_day_income = 0
        if prev_same_day_start and prev_same_day_end:
            prev_same_day_income = (
                sum_tx_income(prev_same_day_start, prev_same_day_end)
                + sum_pay_income(prev_same_day_start, prev_same_day_end)
            )

        # For net profit: get expense transactions separately
        expense_query = Transaction.query.filter(Transaction.category == 'expense')
        if target_user_id:
            expense_query = expense_query.filter(Transaction.admin_id == target_user_id)
        if card_id:
            expense_query = expense_query.filter(Transaction.card_id == card_id)
        if server_id:
            expense_query = expense_query.filter(Transaction.server_id == server_id)
        expense_query = _apply_finance_tx_filters(expense_query, finance_filters, allow_income=False)

        total_expense = db.session.query(func.coalesce(func.sum(Transaction.amount), 0)).filter(
            Transaction.id.in_(expense_query.with_entities(Transaction.id))
        ).scalar() or 0

        month_expense = db.session.query(func.coalesce(func.sum(Transaction.amount), 0)).filter(
            Transaction.id.in_(expense_query.with_entities(Transaction.id)),
            Transaction.created_at >= month_start
        ).scalar() or 0

        prev_month_expense = 0
        if prev_month_start and prev_period_end:
            prev_month_expense = db.session.query(func.coalesce(func.sum(Transaction.amount), 0)).filter(
                Transaction.id.in_(expense_query.with_entities(Transaction.id)),
                Transaction.created_at >= prev_month_start,
                Transaction.created_at < prev_period_end
            ).scalar() or 0

        month_cost_abs = abs(month_expense)
        prev_month_cost_abs = abs(prev_month_expense)
        month_profit = int(month_income) - int(month_cost_abs)
        prev_month_profit = int(prev_month_income) - int(prev_month_cost_abs)

        def pct_change(cur, prev):
            try:
                cur = float(cur)
                prev = float(prev)
            except Exception:
                return None
            if prev == 0:
                return 0.0 if cur == 0 else None
            return ((cur - prev) / abs(prev)) * 100.0

        month_income_pct = pct_change(month_income, prev_month_income)
        month_cost_pct = pct_change(month_cost_abs, prev_month_cost_abs)
        month_profit_pct = pct_change(month_profit, prev_month_profit)

        today_income_pct = pct_change(today_income, prev_same_day_income)

        month_income_delta = int(month_income) - int(prev_month_income)
        month_cost_delta = int(month_cost_abs) - int(prev_month_cost_abs)
        month_profit_delta = int(month_profit) - int(prev_month_profit)
        today_income_delta = int(today_income) - int(prev_same_day_income)

        payment_count = tx_income_query.count() + pay_income_query.count()

        income_by_card = []
        if user.is_superadmin or user.role == 'reseller':
            card_stats = db.session.query(
                BankCard.id,
                BankCard.label,
                BankCard.bank_name,
                func.sum(Transaction.amount).label('total')
            ).join(Transaction, Transaction.card_id == BankCard.id).filter(
                Transaction.id.in_(tx_income_query.with_entities(Transaction.id)),
                Transaction.category.in_(['income', 'expense'])
            ).group_by(BankCard.id).all()

            for card_id, label, bank_name, total in card_stats:
                income_by_card.append({
                    'card_id': card_id,
                    'label': label,
                    'bank_name': bank_name,
                    'total': total or 0
                })
        return jsonify({
            'success': True,
            'stats': {
                'today': today_income,
                'month': month_income,
                'month_expense': abs(month_expense),
                'total': total_income,
                'prev_month': prev_month_income,
                'prev_month_expense': prev_month_cost_abs,
                'month_net': month_profit,
                'prev_month_net': prev_month_profit,
                'month_change_pct': month_income_pct,
                'month_expense_change_pct': month_cost_pct,
                'month_net_change_pct': month_profit_pct,
                'today_prev': prev_same_day_income,
                'today_change_pct': today_income_pct,
                'today_change_amount': today_income_delta,
                'month_change_amount': month_income_delta,
                'month_expense_change_amount': month_cost_delta,
                'month_net_change_amount': month_profit_delta,
                'payment_count': payment_count,
                'by_card': income_by_card,
                'total_income': total_income,
                'total_expense': total_expense
            }
        })


def _statement_pkg_from_desc(desc):
    """Best-effort package name from a legacy transaction description (rows
    created before the package_name column existed)."""
    if not desc:
        return None
    s = str(desc)
    for key in ('Purchase Package:', 'Renew Package:'):
        if key in s:
            return (s.split(key, 1)[1].rsplit(' - ', 1)[0].strip() or None)
    if 'Custom' in s:
        return 'Custom'
    return None


def _statement_plan_from_desc(desc):
    """Best-effort (volume_gb, days) from a legacy custom-plan description like
    'Custom Plan: 30 Days, 50 GB - email' or 'Renew Custom: 30 Days, 50 GB - …'.
    Returns (None, None) when not present (e.g. 'Unlimited')."""
    if not desc:
        return (None, None)
    s = str(desc)
    dm = re.search(r'(\d+)\s*Days', s, re.IGNORECASE)
    gm = re.search(r'(\d+)\s*GB', s, re.IGNORECASE)
    return (int(gm.group(1)) if gm else None, int(dm.group(1)) if dm else None)


_STATEMENT_SPEND_TYPES = ('purchase', 'renew', 'reset_traffic')
_STATEMENT_DEPOSIT_TYPES = ('manual_receipt', 'manual_receipt_auto', 'manual_receipt_reversal')
_STATEMENT_GB = float(1024 ** 3)


def _statement_plan_resolver():
    """Resolve a transaction row's effective (volume_gb, days): prefer the per-row
    columns, else the Package definition, else parse the legacy description."""
    try:
        pkg_lookup = {p.name: (int(p.volume or 0), int(p.days or 0)) for p in Package.query.all()}
    except Exception:
        pkg_lookup = {}

    def resolve(r, pname):
        vol, days = r.volume_gb, r.days
        if vol is None or days is None:
            pv = pd = None
            if pname and pname in pkg_lookup:
                pv, pd = pkg_lookup[pname]
            if pv is None or pd is None:
                dv, dd = _statement_plan_from_desc(r.description)
                pv = pv if pv is not None else dv
                pd = pd if pd is not None else dd
            if vol is None:
                vol = pv
            if days is None:
                days = pd
        return int(vol or 0), int(days or 0)

    return resolve


def _statement_access_target(user):
    """Which reseller a statement request targets: a superadmin may pick anyone,
    a reseller is locked to their own figures."""
    if user.role == 'superadmin' or user.is_superadmin:
        target_id = request.args.get('user_id', type=int) or user.id
    else:
        target_id = user.id
    return db.session.get(Admin, target_id)


def _statement_date_range():
    from app import parse_jalali_date  # deferred: app-level helper, avoids circular import
    start_dt = parse_jalali_date(request.args.get('start_date'), end_of_day=False)
    end_dt = parse_jalali_date(request.args.get('end_date'), end_of_day=True)
    return start_dt, end_dt


def _reseller_sub_id_map(target_id, server_id=None):
    """sub_id -> ClientOwnership for every account a reseller owns, optionally
    narrowed to one server. UsageDaily is keyed by the panel's subscription id
    (`subId`), which often differs from the stored client uuid — so the map
    covers both the stored uuid and the live-resolved subId."""
    from app import GLOBAL_SERVER_DATA  # deferred: app-level helper, avoids circular import
    q = ClientOwnership.query.filter_by(reseller_id=target_id)
    if server_id is not None:
        q = q.filter_by(server_id=server_id)
    ownerships = q.all()
    out = {}
    by_uuid = {}
    by_email = {}
    for o in ownerships:
        uuid = (o.client_uuid or '').strip()
        if uuid:
            out[uuid] = o
            by_uuid[uuid.lower()] = o
        email = (o.client_email or '').strip().lower()
        if email:
            by_email.setdefault((o.server_id, email), o)
    # Resolve the live panel subId for each owned account (subId != uuid on
    # many panels), so the UsageDaily rollups actually join.
    for inbound in GLOBAL_SERVER_DATA.get('inbounds') or []:
        try:
            sid = int(inbound.get('server_id'))
        except (TypeError, ValueError):
            continue
        if server_id is not None and sid != server_id:
            continue
        for client in inbound.get('clients') or []:
            sub_id = str(client.get('subId') or client.get('id') or '').strip()
            if not sub_id:
                continue
            cid = (client.get('id') or '').strip().lower()
            email = (client.get('email') or '').strip().lower()
            owner = by_uuid.get(cid) or by_email.get((sid, email))
            if owner is not None:
                out.setdefault(sub_id, owner)
    return out


def _usage_daily_sums(sub_ids, start_dt, end_dt):
    """({sub_id: used_bytes}, matched_row_count) over the range from the
    UsageDaily rollups. IN-lists are chunked so big resellers stay under
    SQLite's variable limit. Row count tells callers whether snapshots even
    exist for the range (0 => usage figures are unknowable, not zero)."""
    from app import _usage_tehran_date  # deferred: app-level helper, avoids circular import
    sub_list = [s for s in sub_ids if s]
    if not sub_list:
        return {}, 0
    from_date = _usage_tehran_date(start_dt) if start_dt else datetime(2000, 1, 1).date()
    to_date = _usage_tehran_date(end_dt) if end_dt else _usage_tehran_date(datetime.utcnow())
    totals = {}
    matched = 0
    for i in range(0, len(sub_list), 500):
        chunk = sub_list[i:i + 500]
        pairs = (db.session.query(UsageDaily.sub_id,
                                  db.func.sum(UsageDaily.upload_bytes + UsageDaily.download_bytes))
                 .filter(UsageDaily.sub_id.in_(chunk),
                         UsageDaily.usage_date >= from_date,
                         UsageDaily.usage_date <= to_date)
                 .group_by(UsageDaily.sub_id).all())
        matched += len(pairs)
        for sub_id, total in pairs:
            if total:
                totals[sub_id] = int(total)
    return totals, matched


def _usage_first_date(sub_ids):
    """Earliest UsageDaily date across these subs (None when no snapshots at
    all) — used to warn that usage before that date is unknowable."""
    sub_list = [s for s in sub_ids if s]
    first = None
    for i in range(0, len(sub_list), 500):
        chunk = sub_list[i:i + 500]
        candidate = (db.session.query(db.func.min(UsageDaily.usage_date))
                     .filter(UsageDaily.sub_id.in_(chunk)).scalar())
        if candidate and (first is None or candidate < first):
            first = candidate
    return first


def _compute_reseller_statement(target, start_dt, end_dt, pct=None):
    """Full statement dataset for one reseller over a naive-UTC range: ledger
    aggregates, charged-vs-used traffic and its cost, pricing mode, and
    settlement figures. `pct` (optional) computes a percentage-based debt on the
    period's spend, priced at what each GB cost when it was bought."""
    from app import _usage_tehran_date, format_jalali  # deferred: app-level helper, avoids circular import
    q = Transaction.query.filter(Transaction.admin_id == target.id)
    if start_dt:
        q = q.filter(Transaction.created_at >= start_dt)
    if end_dt:
        q = q.filter(Transaction.created_at <= end_dt)
    rows = q.order_by(Transaction.created_at.desc()).all()

    resolve = _statement_plan_resolver()

    created = renewed = reset_cnt = 0
    spent = 0
    deposited = 0
    charged_gb = 0.0
    paid_volume_gb = 0
    paid_charge = 0
    by_package = {}
    by_server = {}
    out_rows = []

    for r in rows:
        amt = int(r.amount or 0)
        t = r.type or ''
        row_d = r.to_dict()
        row_d['pkg'] = None
        row_d['is_gift'] = False
        row_d['charge'] = 0
        if t in _STATEMENT_SPEND_TYPES:
            charge = -amt if amt < 0 else 0  # reseller credit usage is stored negative
            spent += charge
            row_d['charge'] = charge
            pname = r.package_name or _statement_pkg_from_desc(r.description)
            eff_vol, eff_days = resolve(r, pname)
            is_gift = (amt == 0) or ('(Free)' in (r.description or ''))
            charged_gb += eff_vol
            if not is_gift and eff_vol > 0:
                paid_volume_gb += eff_vol
                paid_charge += charge
            if t == 'purchase':
                created += 1
            elif t == 'renew':
                renewed += 1
            else:
                reset_cnt += 1
            if t in ('purchase', 'renew'):
                pname = pname or '—'
                row_d['pkg'] = pname
                row_d['eff_volume_gb'] = eff_vol
                row_d['eff_days'] = eff_days
                row_d['is_gift'] = is_gift
                b = by_package.setdefault(pname, {'count': 0, 'spent': 0, 'volume_gb': 0, 'days': 0, 'gifts': 0, 'items': []})
                b['count'] += 1
                b['spent'] += charge
                b['volume_gb'] += eff_vol
                b['days'] += eff_days
                if is_gift:
                    b['gifts'] += 1
                # Embed the per-user line items so the UI drill-down doesn't have to
                # re-match anything client-side.
                b['items'].append({
                    'email': r.client_email or '—',
                    'vol': eff_vol, 'days': eff_days,
                    'amount': charge, 'gift': is_gift, 'type': t,
                    'date': format_jalali(r.created_at),
                })
            skey = (r.server_id or 0, (r.server.name if r.server else None) or '—')
            sb = by_server.setdefault(skey, {'count': 0, 'spent': 0})
            sb['count'] += 1
            sb['spent'] += charge
        elif t in _STATEMENT_DEPOSIT_TYPES:
            deposited += amt  # deposits positive, reversals negative
        row_d['date_tehran'] = _usage_tehran_date(r.created_at).isoformat() if r.created_at else None
        out_rows.append(row_d)

    # Charged vs actually-consumed traffic, from the daily usage rollups.
    sub_map = _reseller_sub_id_map(target.id)
    usage, usage_rows = _usage_daily_sums(sub_map.keys(), start_dt, end_dt)
    used_bytes = sum(usage.values())
    used_gb = round(used_bytes / _STATEMENT_GB, 2)
    remaining_gb = round(max(charged_gb - used_gb, 0), 2)
    if paid_volume_gb > 0:
        # Weighted effective price: what a GB actually cost in this period's charges.
        price_per_gb = int(round(paid_charge / float(paid_volume_gb)))
        price_basis = 'period_transactions'
    elif target.custom_cost_per_gb:
        price_per_gb = int(target.custom_cost_per_gb)
        price_basis = 'reseller_fixed_rate'
    else:
        price_per_gb = 0
        price_basis = 'unknown'
    usage_cost = int(round((used_bytes / _STATEMENT_GB) * price_per_gb)) if price_per_gb else 0

    # Snapshot coverage: no rollup rows in range means "unknown", not "zero".
    usage_available = usage_rows > 0 or not sub_map
    coverage_from_jalali = None
    if not usage_available:
        first = _usage_first_date(sub_map.keys())
        if first:
            coverage_from_jalali = format_jalali(datetime.combine(first, datetime.min.time()))

    if target.custom_cost_per_gb or target.custom_cost_per_day:
        pricing_mode = 'fixed'
    elif target.discount_percent:
        pricing_mode = 'percentage'
    else:
        pricing_mode = 'standard'

    balance = deposited - spent  # negative => under-deposited (debt)
    summary = {
        'reseller': {'id': target.id, 'username': target.username,
                     'role': target.role, 'credit': target.credit or 0},
        'created': created, 'renewed': renewed, 'reset': reset_cnt,
        'spent': spent, 'deposited': deposited,
        'should_deposit': spent, 'balance': balance,
        'row_count': len(rows),
        'start_jalali': format_jalali(start_dt) if start_dt else None,
        'end_jalali': format_jalali(end_dt) if end_dt else None,
        'revenue': deposited,   # what the reseller paid us in the period
        'should_pay': spent,    # what the period's charges cost them
        'traffic': {
            'charged_gb': round(charged_gb, 2), 'used_gb': used_gb,
            'remaining_gb': remaining_gb, 'usage_cost': usage_cost,
            'price_per_gb': price_per_gb, 'price_basis': price_basis,
            'accounts_tracked': len({o.id for o in sub_map.values()}),
            'usage_available': usage_available,
            'coverage_from_jalali': coverage_from_jalali,
        },
        'pricing': {
            'mode': pricing_mode,
            'discount_percent': target.discount_percent or 0,
            'custom_cost_per_gb': target.custom_cost_per_gb,
            'custom_cost_per_day': target.custom_cost_per_day,
        },
    }
    if pct is not None:
        summary['percent_settlement'] = {
            'percent': pct, 'base': spent,
            'debt': int(round(spent * pct / 100.0)),
        }
    by_package_list = sorted(({'package': k, **v} for k, v in by_package.items()),
                             key=lambda x: x['spent'], reverse=True)
    by_server_list = sorted(({'server_id': k[0] or None, 'server': k[1], **v}
                             for k, v in by_server.items()),
                            key=lambda x: x['spent'], reverse=True)
    return {'summary': summary, 'by_package': by_package_list,
            'by_server': by_server_list, 'rows': out_rows}


def _reseller_accounts_breakdown(target, start_dt, end_dt, server_id=None):
    """Per-account statement lines: charged GB/amount in range, used GB in range,
    remaining, live panel status, and the last change. server_id=None => all
    servers; the statement UI calls this lazily per server so big resellers
    stay responsive."""
    from app import GLOBAL_SERVER_DATA, format_jalali  # deferred: app-level helper, avoids circular import
    sub_map = _reseller_sub_id_map(target.id, server_id)
    usage, _usage_rows = _usage_daily_sums(sub_map.keys(), start_dt, end_dt)
    resolve = _statement_plan_resolver()

    # Live status from the panel cache.
    live = {}
    for inbound in GLOBAL_SERVER_DATA.get('inbounds') or []:
        try:
            sid = int(inbound.get('server_id'))
        except (TypeError, ValueError):
            continue
        if server_id is not None and sid != server_id:
            continue
        for client in inbound.get('clients') or []:
            sub_id = str(client.get('subId') or client.get('id') or '').strip()
            if sub_id and sub_id in sub_map:
                live[sub_id] = 'active' if client.get('enable', True) else 'disabled'

    # Charged volume/amount per account from the range's spend rows.
    q = Transaction.query.filter(Transaction.admin_id == target.id,
                                 Transaction.type.in_(_STATEMENT_SPEND_TYPES))
    if server_id is not None:
        q = q.filter(Transaction.server_id == server_id)
    if start_dt:
        q = q.filter(Transaction.created_at >= start_dt)
    if end_dt:
        q = q.filter(Transaction.created_at <= end_dt)
    per_email = {}
    for r in q.order_by(Transaction.created_at.desc()).all():
        email = (r.client_email or '').strip().lower()
        if not email:
            continue
        charge = -int(r.amount or 0) if int(r.amount or 0) < 0 else 0
        pname = r.package_name or _statement_pkg_from_desc(r.description)
        vol, _days = resolve(r, pname)
        a = per_email.setdefault((r.server_id or 0, email),
                                 {'charged_gb': 0, 'spent': 0, 'events': 0,
                                  'last_type': None, 'last_at': None})
        a['charged_gb'] += vol
        a['spent'] += charge
        a['events'] += 1
        if a['last_at'] is None:  # rows arrive newest-first
            a['last_at'] = r.created_at
            a['last_type'] = r.type

    server_names = {s.id: s.name for s in Server.query.all()}
    # One ownership can map to several sub_ids (stored uuid + live panel subId);
    # aggregate per ownership so accounts never appear twice.
    owner_subs = {}
    for sub_id, o in sub_map.items():
        entry = owner_subs.setdefault(o.id, {'ownership': o, 'sub_ids': set()})
        entry['sub_ids'].add(sub_id)
    accounts = []
    for entry in owner_subs.values():
        o = entry['ownership']
        sub_ids = entry['sub_ids']
        agg = per_email.get((o.server_id or 0, (o.client_email or '').strip().lower()))
        used_bytes = sum(usage.get(s, 0) for s in sub_ids)
        used_gb = round(used_bytes / _STATEMENT_GB, 2)
        charged = agg['charged_gb'] if agg else 0
        status = next((live[s] for s in sub_ids if s in live), 'deleted')
        accounts.append({
            'server_id': o.server_id,
            'server': server_names.get(o.server_id, '—'),
            'email': o.client_email,
            'status': status,
            'charged_gb': charged,
            'used_gb': used_gb,
            'remaining_gb': round(max(charged - used_gb, 0), 2),
            'spent': agg['spent'] if agg else 0,
            'events': agg['events'] if agg else 0,
            'last_type': agg['last_type'] if agg else None,
            'last_change': format_jalali(agg['last_at']) if agg and agg['last_at'] else None,
        })
    accounts.sort(key=lambda a: (a['spent'], a['used_gb']), reverse=True)
    return accounts


@bp.route('/api/finance/reseller-statement', methods=['GET'])
@login_required
def reseller_statement():
    """Per-reseller accounting statement over a Jalali date range: accounts
    created / renewed / reset, charged vs used traffic and its cost, pricing
    mode, optional percentage settlement, and deposited vs should-deposit.
    Superadmin can pick any reseller; a reseller only ever sees their own figures."""
    user = db.session.get(Admin, session['admin_id'])
    if not user:
        return jsonify({'success': False, 'error': 'User not found'}), 401
    target = _statement_access_target(user)
    if not target:
        return jsonify({'success': False, 'error': 'Reseller not found'}), 404
    start_dt, end_dt = _statement_date_range()
    pct = request.args.get('percent', type=float)
    data = _compute_reseller_statement(target, start_dt, end_dt, pct)
    resp = jsonify({'success': True, **data})
    resp.headers['Cache-Control'] = 'no-store, max-age=0'
    return resp


@bp.route('/api/finance/reseller-statement/accounts', methods=['GET'])
@login_required
def reseller_statement_accounts():
    """Lazy per-server account drill-down for the statement: which accounts the
    reseller has on one server, with charged/used/remaining GB and last change."""
    user = db.session.get(Admin, session['admin_id'])
    if not user:
        return jsonify({'success': False, 'error': 'User not found'}), 401
    target = _statement_access_target(user)
    if not target:
        return jsonify({'success': False, 'error': 'Reseller not found'}), 404
    server_id = request.args.get('server_id', type=int)
    start_dt, end_dt = _statement_date_range()
    accounts = _reseller_accounts_breakdown(target, start_dt, end_dt, server_id=server_id)
    resp = jsonify({'success': True, 'accounts': accounts})
    resp.headers['Cache-Control'] = 'no-store, max-age=0'
    return resp


@bp.route('/api/finance/reseller-statement/export', methods=['GET'])
@login_required
def reseller_statement_export():
    """Full statement workbook (xlsx): summary, per-day charge feed, per-account
    usage, and the raw ledger rows for the selected range."""
    user = db.session.get(Admin, session['admin_id'])
    if not user:
        return jsonify({'success': False, 'error': 'User not found'}), 401
    target = _statement_access_target(user)
    if not target:
        return jsonify({'success': False, 'error': 'Reseller not found'}), 404
    start_dt, end_dt = _statement_date_range()
    pct = request.args.get('percent', type=float)
    data = _compute_reseller_statement(target, start_dt, end_dt, pct)
    accounts = _reseller_accounts_breakdown(target, start_dt, end_dt)

    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({'success': False, 'error': 'openpyxl is not installed'}), 500

    s = data['summary']
    traffic = s['traffic']
    wb = Workbook(write_only=True)

    ws = wb.create_sheet('Summary')
    ws.append(['Reseller Statement'])
    ws.append(['Reseller', s['reseller']['username']])
    ws.append(['Period', '{} → {}'.format(s.get('start_jalali') or '…', s.get('end_jalali') or 'today')])
    ws.append([])
    ws.append(['Metric', 'Value'])
    ws.append(['Accounts created', s['created']])
    ws.append(['Accounts renewed', s['renewed']])
    ws.append(['Traffic resets', s['reset']])
    ws.append(['Charged volume (GB)', traffic['charged_gb']])
    ws.append(['Used volume (GB)', traffic['used_gb']])
    ws.append(['Remaining volume (GB)', traffic['remaining_gb']])
    ws.append(['Effective price per GB (T)', traffic['price_per_gb']])
    ws.append(['Price basis', traffic['price_basis']])
    ws.append(['Usage cost (T)', traffic['usage_cost']])
    ws.append(['Pricing mode', s['pricing']['mode']])
    ws.append(['Discount percent', s['pricing']['discount_percent']])
    ws.append(['Should pay (T)', s['should_pay']])
    ws.append(['Revenue / deposited (T)', s['revenue']])
    ws.append(['Balance (T)', s['balance']])
    if not traffic.get('usage_available', True):
        note = 'No usage snapshots in this period — Used/Remaining GB and Usage cost are understated.'
        if traffic.get('coverage_from_jalali'):
            note += ' Earliest snapshot: {}.'.format(traffic['coverage_from_jalali'])
        ws.append(['⚠ Usage data', note])
    ps = s.get('percent_settlement')
    if ps:
        ws.append(['Debt at {}% (T)'.format(ps['percent']), ps['debt']])

    # Per-day charge feed: which users were topped up each day, and the payable.
    ws = wb.create_sheet('Daily')
    ws.append(['Date', 'User', 'Type', 'Package', 'GB', 'Days',
               'Amount (T)', 'Payable (T)', 'Gift', 'Server'])
    daily = {}
    for r in data['rows']:
        if r.get('type') not in _STATEMENT_SPEND_TYPES:
            continue
        dkey = r.get('date_tehran') or (r.get('date') or '')[:10]
        daily.setdefault(dkey, []).append(r)
    for dkey in sorted(daily):
        day_gb = day_amt = day_pay = 0
        for r in sorted(daily[dkey], key=lambda x: x.get('date') or ''):
            amt = int(r.get('charge') or 0)
            payable = int(round(amt * pct / 100.0)) if pct is not None else amt
            vol = r.get('eff_volume_gb') if r.get('eff_volume_gb') is not None else (r.get('volume_gb') or 0)
            dys = r.get('eff_days') if r.get('eff_days') is not None else (r.get('days') or 0)
            day_gb += vol or 0
            day_amt += amt
            day_pay += payable
            ws.append([dkey, r.get('client_email') or '—', r.get('type') or '',
                       r.get('pkg') or r.get('package_name') or '',
                       vol or 0, dys or 0, amt, payable,
                       'yes' if r.get('is_gift') else '',
                       (r.get('server') or {}).get('name') or ''])
        ws.append([dkey, '— Daily subtotal —', '', '', day_gb, '', day_amt, day_pay, '', ''])

    ws = wb.create_sheet('Accounts')
    ws.append(['Server', 'User', 'Status', 'Charged GB', 'Used GB',
               'Remaining GB', 'Spent (T)', 'Events', 'Last change', 'Last type'])
    for a in accounts:
        ws.append([a['server'], a['email'], a['status'], a['charged_gb'], a['used_gb'],
                   a['remaining_gb'], a['spent'], a['events'],
                   a['last_change'] or '', a['last_type'] or ''])

    ws = wb.create_sheet('Transactions')
    ws.append(['Date', 'Type', 'User', 'Package', 'GB', 'Days',
               'Amount (T)', 'Gift', 'Server', 'Description'])
    for r in sorted(data['rows'], key=lambda x: x.get('date') or ''):
        ws.append([r.get('date_jalali') or '', r.get('type') or '',
                   r.get('client_email') or '',
                   r.get('pkg') or r.get('package_name') or '',
                   r.get('eff_volume_gb') if r.get('eff_volume_gb') is not None else (r.get('volume_gb') or ''),
                   r.get('eff_days') if r.get('eff_days') is not None else (r.get('days') or ''),
                   int(r.get('amount') or 0), 'yes' if r.get('is_gift') else '',
                   (r.get('server') or {}).get('name') or '',
                   r.get('description') or ''])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    who = re.sub(r'\s+', '_', target.username or 'reseller')
    rng = '{}_{}'.format(start_dt.date().isoformat() if start_dt else 'all',
                         end_dt.date().isoformat() if end_dt else 'now')
    return send_file(buf, as_attachment=True,
                     download_name='statement_{}_{}.xlsx'.format(who, rng),
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/api/finance/overview', methods=['GET'])
@login_required
def get_finance_overview():
    """Timeseries overview for income/expense/profit."""
    from app import format_jalali  # deferred: app-level helper, avoids circular import
    user = db.session.get(Admin, session['admin_id'])
    if not user:
        return jsonify({"success": False, "error": "User not found"}), 401

    requested_range = (request.args.get('range') or '30d').strip().lower()
    if requested_range not in ('12m', '30d', '7d', '24h', 'month'):
        requested_range = '30d'

    # Optional month selection (Jalali)
    selected_month = request.args.get('month', type=int)
    selected_year = request.args.get('year', type=int)
    if selected_month and 1 <= int(selected_month) <= 12:
        requested_range = 'month'

    card_id = request.args.get('card_id', type=int)
    server_id = request.args.get('server_id', type=int)
    target_user_id = request.args.get('user_id', type=int)
    finance_filters = _finance_request_filters()
    # The overview has its own Jalali month/window selector. Date chips on the
    # table should not collapse that month, but search/type/direction still need
    # to match the visible Finance context.
    finance_filters['start_dt'] = None
    finance_filters['end_dt'] = None

    tehran_offset = timedelta(hours=3, minutes=30)
    now_utc = datetime.utcnow()
    now_tehran = now_utc + tehran_offset

    def as_tehran(dt_utc):
        return (dt_utc + tehran_offset) if dt_utc else None

    labels = []
    keys = []
    start_utc = None
    end_utc = None

    excluded_types_upper = {'SERVER_COST', 'SERVER_RENEWAL', 'SERVER_TRAFFIC'}

    # Build bucket keys (chronological)
    if requested_range == '12m':
        j_now = jdatetime_class.fromgregorian(datetime=now_tehran)
        j_year = int(j_now.year)
        j_month = int(j_now.month)

        def shift_month(year, month, delta):
            total = (year * 12 + (month - 1)) + delta
            new_year = total // 12
            new_month = (total % 12) + 1
            return new_year, new_month

        month_starts_tehran = []
        for i in range(11, -1, -1):
            y, m = shift_month(j_year, j_month, -i)
            j_start = jdatetime_class(y, m, 1, 0, 0, 0, 0)
            g_start_tehran = j_start.togregorian()
            month_starts_tehran.append(g_start_tehran)
            labels.append(f"{y:04d}/{m:02d}")
            keys.append((y, m))

        # end is start of next month
        next_y, next_m = shift_month(j_year, j_month, 1)
        j_end = jdatetime_class(next_y, next_m, 1, 0, 0, 0, 0)
        g_end_tehran = j_end.togregorian()
        start_utc = month_starts_tehran[0] - tehran_offset
        end_utc = g_end_tehran - tehran_offset
    elif requested_range in ('30d', '7d'):
        days = 30 if requested_range == '30d' else 7
        end_tehran = now_tehran
        end_tehran_floor = end_tehran.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)
        start_tehran = end_tehran_floor - timedelta(days=days)
        start_utc = start_tehran - tehran_offset
        end_utc = end_tehran_floor - tehran_offset

        for i in range(days, 0, -1):
            day_start_tehran = end_tehran_floor - timedelta(days=i)
            # label as Jalali date
            j_label = (format_jalali(day_start_tehran) or '').strip()
            if ' ' in j_label:
                j_label = j_label.split(' ')[0]
            labels.append(j_label)
            keys.append(day_start_tehran.date())
    elif requested_range == 'month':
        # Daily buckets for a selected Jalali month (defaults to current Jalali year)
        j_now = jdatetime_class.fromgregorian(datetime=now_tehran)
        j_year = int(selected_year or j_now.year)
        j_month = int(selected_month or j_now.month)

        # Start of selected month
        j_start = jdatetime_class(j_year, j_month, 1, 0, 0, 0, 0)
        g_start_tehran = j_start.togregorian()
        start_utc = g_start_tehran - tehran_offset

        # Start of next month
        if j_month == 12:
            j_end = jdatetime_class(j_year + 1, 1, 1, 0, 0, 0, 0)
        else:
            j_end = jdatetime_class(j_year, j_month + 1, 1, 0, 0, 0, 0)
        g_end_tehran = j_end.togregorian()
        end_utc = g_end_tehran - tehran_offset

        # Number of days in this Jalali month
        try:
            days_in_month = int(j_end.togregorian().date() - j_start.togregorian().date()).days
        except Exception:
            days_in_month = 31
            if j_month > 6:
                days_in_month = 30
            if j_month == 12:
                days_in_month = 29

        for day in range(1, days_in_month + 1):
            # Use day number as label (cleaner for bar chart)
            labels.append(str(day))
            # Keys are Gregorian dates in Tehran (date-only)
            g_day_tehran = (jdatetime_class(j_year, j_month, day, 0, 0, 0, 0)).togregorian()
            keys.append(g_day_tehran.date())
    else:  # 24h
        end_tehran = now_tehran.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
        start_tehran = end_tehran - timedelta(hours=24)
        start_utc = start_tehran - tehran_offset
        end_utc = end_tehran - tehran_offset

        for i in range(24, 0, -1):
            hour_start_tehran = end_tehran - timedelta(hours=i)
            labels.append(hour_start_tehran.strftime('%H:00'))
            keys.append(hour_start_tehran)

    # Query ledger rows in a single window then bin in Python.
    # - Income: any positive amount EXCEPT server-cost/renewal/traffic.
    # - Expense: any negative amount (absolute).
    # Also include manual `Payment` rows as income.
    tx_query = Transaction.query
    if user.role == 'reseller':
        tx_query = tx_query.filter(Transaction.admin_id == user.id)
    else:
        if target_user_id:
            tx_query = tx_query.filter(Transaction.admin_id == target_user_id)

    if card_id:
        tx_query = tx_query.filter(Transaction.card_id == card_id)
    if server_id:
        tx_query = tx_query.filter(Transaction.server_id == server_id)

    tx_query = tx_query.filter(Transaction.created_at >= start_utc, Transaction.created_at < end_utc)
    tx_query = tx_query.filter(Transaction.type != 'audit')
    tx_query = _apply_finance_tx_filters(tx_query, finance_filters)

    tx_rows = tx_query.all()

    pay_query = Payment.query.filter(Payment.payment_date >= start_utc, Payment.payment_date < end_utc)
    if user.role == 'reseller':
        pay_query = pay_query.filter(Payment.admin_id == user.id)
    else:
        if target_user_id:
            pay_query = pay_query.filter(Payment.admin_id == target_user_id)
    if card_id:
        pay_query = pay_query.filter(Payment.card_id == card_id)
    pay_query = _apply_finance_payment_filters(pay_query, finance_filters)
    pay_rows = pay_query.all()

    income_map = {k: 0 for k in keys}
    expense_map = {k: 0 for k in keys}

    def bucket_for_tehran_dt(tehran_dt):
        if requested_range == '12m':
            j_dt = jdatetime_class.fromgregorian(datetime=tehran_dt)
            return (int(j_dt.year), int(j_dt.month))
        if requested_range in ('30d', '7d', 'month'):
            return tehran_dt.date()
        return tehran_dt.replace(minute=0, second=0, microsecond=0)

    for tx in tx_rows:
        if not tx.created_at:
            continue
        tehran_dt = as_tehran(tx.created_at)
        if not tehran_dt:
            continue

        bucket_key = bucket_for_tehran_dt(tehran_dt)
        if bucket_key not in income_map:
            continue

        amount = int(tx.amount or 0)
        if amount > 0 and finance_filters.get('direction') != 'expense':
            tx_type = (tx.type or '')
            if tx_type and str(tx_type).upper() in excluded_types_upper:
                continue
            income_map[bucket_key] += amount
        elif amount < 0 and finance_filters.get('direction') != 'income':
            expense_map[bucket_key] += abs(amount)

    for pay in pay_rows:
        if not pay.payment_date:
            continue
        tehran_dt = as_tehran(pay.payment_date)
        if not tehran_dt:
            continue

        bucket_key = bucket_for_tehran_dt(tehran_dt)
        if bucket_key not in income_map:
            continue

        amount = int(pay.amount or 0)
        if amount > 0 and finance_filters.get('direction') != 'expense':
            income_map[bucket_key] += amount

    income_series = [int(income_map[k] or 0) for k in keys]
    expense_series = [int(expense_map[k] or 0) for k in keys]
    profit_series = [int(i - e) for i, e in zip(income_series, expense_series)]

    return jsonify({
        'success': True,
        'range': requested_range,
        'month': int(selected_month) if selected_month else None,
        'year': int(selected_year) if selected_year else None,
        'labels': labels,
        'series': {
            'income': income_series,
            'expense': expense_series,
            'profit': profit_series
        }
    })
@bp.route('/api/customers/<int:customer_id>/credit', methods=['POST'])
@limiter.limit('20 per minute')
@login_required
def adjust_customer_credit(customer_id):
    from app import _log_audit, _telegram_operations_admin  # deferred: app-level helper, avoids circular import
    admin = _telegram_operations_admin()
    if not admin or not (admin.is_superadmin or str(admin.role or '').lower() in ('admin', 'superadmin')):
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    customer = db.session.get(CustomerAccount, customer_id)
    if not customer:
        return jsonify({'success': False, 'error': 'Customer not found'}), 404
    payload = request.get_json(silent=True) or {}
    try:
        amount = int(payload.get('amount') or 0)
    except (TypeError, ValueError):
        amount = 0
    if amount == 0:
        return jsonify({'success': False, 'error': 'A non-zero amount is required'}), 400
    customer.credit = int(customer.credit or 0) + amount
    db.session.add(CustomerTransaction(
        customer_id=customer.id,
        type='adjust',
        amount=amount,
        description=str(payload.get('description') or '').strip()[:255] or None,
    ))
    _log_audit('customer.credit_adjust', customer, actor=admin, meta={'amount': amount})
    db.session.commit()
    return jsonify({'success': True, 'credit': int(customer.credit or 0)})
